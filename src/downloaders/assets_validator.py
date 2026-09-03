"""
Validator for the manual Global Energy Monitor (GEM) snapshot.

DOWNLOADS NOTHING. Unlike the other modules in this package, this one is
read-only over an ``.xlsx`` you download by hand from the GEM Global
Integrated Power Tracker and place in ``data/raw/assets/``. The snapshot must
be versioned and dated in the manuscript.

AGGREGATION AND FUEL BUCKETING ONLY — no index, weighting or resilience
logic. Pipeline, in order:

1. Split by ``Status``: ``operating`` continues to the main pipeline;
   ``construction`` / ``announced`` / ``pre-construction`` are preserved
   separately (all countries) in ``gem_planned_assets.csv``; every other
   status is counted and reported, never silently dropped.
2. Filter to the three study countries (done before aggregation — the
   grouping key includes country, so the result is identical and far
   cheaper).
3. Write the per-unit detail (GEM publishes one row per generating unit) to
   ``gem_units_detail.csv``.
4. Aggregate to plants: key = country + normalised name + coordinate rounded
   to ``DUPLICATE_COORD_TOLERANCE_DEG`` (~100 m). ``capacity_mw`` is summed;
   ``commissioning_year`` is the minimum across units (oldest unit);
   divergent ``fuel_type`` across units yields ``None`` +
   ``mixed_fuel_type = True``.
5. Add ``fuel_type_bucket`` in {hydro, wind, solar, thermal}. Coal and every
   other thermal technology are merged into ``thermal`` (ARCHITECTURE.md
   Section 6). Mixed-fuel plants use a per-name override.
6. For countries in ``MAINLAND_ONLY_COUNTRIES`` (Portugal today), physically
   remove plants outside the mainland bounding box and save them, with a
   reason, to ``gem_excluded_azores_madeira.csv``. For other countries,
   out-of-bbox is only reported (it signals GADM boundary imprecision, not a
   scope decision) and the plant stays.

If an essential column is missing, the script raises with a clear message —
it never proceeds filling with NaN.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from src.config import (
    ASSETS_PROCESSED,
    ASSETS_RAW,
    COUNTRIES,
    MAINLAND_ONLY_COUNTRIES,
    OUTPUT_INSPECTION,
)
from src.downloaders.boundaries_downloader import get_country_bounds

logger = logging.getLogger(__name__)

REQUIRED_INTERNAL_COLUMNS = [
    "plant_name", "lat", "lon", "capacity_mw",
    "fuel_type", "commissioning_year", "country",
]

# Real GEM column name -> internal name. Confirmed from the "Power facilities"
# sheet of the Global Integrated Power Tracker export.
COLUMN_MAPPING: dict[str, str] = {
    "Plant / Project name": "plant_name",
    "Latitude": "lat",
    "Longitude": "lon",
    "Capacity (MW)": "capacity_mw",
    "Type": "fuel_type",
    "Start year": "commissioning_year",
    "Country/area": "country",
}

STATUS_COLUMN = "Status"
UNIT_NAME_COLUMN = "Unit / Phase name"

OPERATING_STATUS = "operating"
PLANNED_STATUSES = {"construction", "announced", "pre-construction"}

# ~100 m in decimal degrees at the equator. Open asset registries carry
# rounding noise in the 3rd/4th decimal between units of the same plant, so
# coordinates are rounded to an explicit tolerance rather than compared for
# exact equality.
DUPLICATE_COORD_TOLERANCE_DEG = 0.0009

ASSET_FILE_EXTENSIONS = (".csv", ".xlsx", ".xls")

# Fuel bucket: four categories. Coal and every other thermal technology are
# merged into "thermal" — same physical mechanism (cooling-water dependence
# and temperature sensitivity of that water); the age-curve tension inside the
# merged bucket is handled by verification item V1, not here.
FUEL_TYPE_TO_BUCKET = {
    "hydropower": "hydro",
    "wind": "wind",
    "utility-scale solar": "solar",
    "coal": "thermal",
    "oil/gas": "thermal",
    "nuclear": "thermal",
    "bioenergy": "thermal",
    "geothermal": "thermal",
}

# Per-name overrides for mixed-fuel plants (no single original fuel_type).
# Every known mixed plant is a fossil combination and maps to "thermal";
# the explicit list keeps the set auditable and makes an unlisted mixed
# plant a hard error rather than a silent gap.
MIXED_FUEL_BUCKET_OVERRIDES = {
    "Itaqui power station": "thermal",
    "Vizag Steel Plant power station": "thermal",
    "Hazira power station (Reliance)": "thermal",
    "Guarani power station": "thermal",
    "Figueira da Foz Industrial Complex power station": "thermal",
    "Viana Mill power station": "thermal",
    "Vitória power station": "thermal",
}

MAINLAND_EXCLUSION_FILENAMES = {
    "Portugal": "gem_excluded_azores_madeira.csv",
}


# --------------------------------------------------------------------------
# File discovery / reading
# --------------------------------------------------------------------------
def find_gem_file() -> Path:
    """Locate the GEM snapshot in ``data/raw/assets/`` without assuming a
    file name. Accepts ``.csv``/``.xlsx``/``.xls``. Raises if there are zero
    or more than one candidate — that choice is the user's, not the
    script's."""
    ASSETS_RAW.mkdir(parents=True, exist_ok=True)
    candidates = sorted(
        p for p in ASSETS_RAW.iterdir()
        if p.is_file() and p.suffix.lower() in ASSET_FILE_EXTENSIONS
    )
    if not candidates:
        raise FileNotFoundError(
            f"No .csv/.xlsx/.xls file in {ASSETS_RAW}. Download the GEM Global "
            f"Integrated Power Tracker snapshot by hand and place it there."
        )
    if len(candidates) > 1:
        names = ", ".join(f.name for f in candidates)
        raise FileNotFoundError(
            f"More than one file in {ASSETS_RAW}: {names}. Leave only the GEM "
            f"snapshot to validate."
        )
    return candidates[0]


def _select_data_sheet(wb: openpyxl.workbook.workbook.Workbook) -> str:
    """GEM ``.xlsx`` exports have several sheets (a cover 'About' sheet, the
    data sheet, a reference table). Pick the sheet with the most header
    columns — the data sheet has dozens, the others have one. Logged, never
    silent."""
    if len(wb.sheetnames) == 1:
        return wb.sheetnames[0]

    n_cols = {}
    for name in wb.sheetnames:
        header = next(wb[name].iter_rows(max_row=1, values_only=True), ())
        n_cols[name] = sum(1 for v in header if v is not None)

    chosen = max(n_cols, key=n_cols.get)
    logger.info("Workbook sheets %s; selected data sheet '%s'.", list(n_cols.items()), chosen)
    return chosen


def _read_excel_table(path: Path, max_rows: int | None = None) -> pd.DataFrame:
    """Read ``.xlsx``/``.xls`` via ``openpyxl`` in read-only + ``iter_rows``
    mode. The real GEM export is ~182k rows x ~52 columns; ``pd.read_excel``
    materialises the whole sheet before applying ``nrows`` and is far too
    slow. Read-only + ``iter_rows`` returns the same data in seconds."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_select_data_sheet(wb)]
        row_iter = ws.iter_rows(values_only=True)
        header = list(next(row_iter))
        rows = []
        for i, row in enumerate(row_iter):
            rows.append(row)
            if max_rows is not None and i + 1 >= max_rows:
                break
        return pd.DataFrame(rows, columns=header)
    finally:
        wb.close()


def read_asset_table(path: Path, nrows: int | None = None) -> pd.DataFrame:
    """Read ``.csv`` or ``.xlsx``/``.xls`` transparently."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        return _read_excel_table(path, max_rows=nrows)
    return pd.read_csv(path, nrows=nrows)


def discover_schema(path: Path, n_sample: int = 3) -> None:
    """Print columns, inferred dtypes and a sample — validates nothing.
    Use this to confirm ``COLUMN_MAPPING`` against a fresh export."""
    df = read_asset_table(path, nrows=5000)
    print(f"File: {path.name}  ({path.stat().st_size / 1024:.1f} KB)")
    print(f"Columns: {len(df.columns)}")
    for col in df.columns:
        example = df[col].dropna().iloc[0] if df[col].notna().any() else "N/A"
        print(f"  {col!r:45s} dtype={str(df[col].dtype):10s} example={example!r}")
    print(f"\nFirst {n_sample} rows:")
    with pd.option_context("display.max_columns", None, "display.width", 200):
        print(df.head(n_sample).to_string())


# --------------------------------------------------------------------------
# Pipeline
# --------------------------------------------------------------------------
def _apply_mapping(df: pd.DataFrame) -> pd.DataFrame:
    """Rename real -> internal columns. Raise if a mapped column is absent or
    an essential internal column is still missing after the rename."""
    missing_in_file = [real for real in COLUMN_MAPPING if real not in df.columns]
    if missing_in_file:
        raise ValueError(
            f"COLUMN_MAPPING refers to columns absent from the file: "
            f"{missing_in_file}. Re-run discover_schema — the export may have changed."
        )
    df = df.rename(columns=COLUMN_MAPPING)
    missing_required = [c for c in REQUIRED_INTERNAL_COLUMNS if c not in df.columns]
    if missing_required:
        raise ValueError(f"Essential column(s) missing after mapping: {missing_required}.")
    return df


def _require_raw_column(df: pd.DataFrame, col_name: str, purpose: str) -> None:
    if col_name not in df.columns:
        raise ValueError(
            f"Column '{col_name}' ({purpose}) not found. Available: {list(df.columns)}."
        )


def split_by_status(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """Separate operating plants (main pipeline) from planned ones
    (construction/announced/pre-construction — preserved separately). Every
    other status is counted, not dropped silently."""
    _require_raw_column(df, STATUS_COLUMN, "asset status")
    status_norm = df[STATUS_COLUMN].astype(str).str.strip().str.lower()

    operating = df[status_norm == OPERATING_STATUS].copy()
    planned = df[status_norm.isin(PLANNED_STATUSES)].copy()
    n_other = int((~status_norm.isin({OPERATING_STATUS} | PLANNED_STATUSES)).sum())

    summary = {
        "statuses_found_in_file": sorted(status_norm.unique().tolist()),
        "n_operating": len(operating),
        "n_planned": len(planned),
        "n_other_excluded_from_both": n_other,
    }
    logger.info(
        "Status: operating=%d planned=%d other=%d",
        len(operating), len(planned), n_other,
    )
    return operating, planned, summary


def _round_coord(series: pd.Series) -> pd.Series:
    return (series / DUPLICATE_COORD_TOLERANCE_DEG).round() * DUPLICATE_COORD_TOLERANCE_DEG


def aggregate_by_plant(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate generating units into plants.

    Key: country + normalised plant name + coordinate rounded to ~100 m.
    - ``capacity_mw``: sum with ``min_count=1`` (a plant whose units all lack
      capacity becomes NaN, not 0.0 — "unknown" is not "zero").
    - ``lat``/``lon``: mean of the units.
    - ``commissioning_year``: minimum across units (oldest unit) — conservative
      for a downstream age factor.
    - ``fuel_type``: kept if the units agree; otherwise ``None`` +
      ``mixed_fuel_type = True`` + ``fuel_types_found`` listing every value.
    - ``n_units``: audit metadata.
    """
    work = df.copy()
    work["_lat_r"] = _round_coord(pd.to_numeric(work["lat"], errors="coerce"))
    work["_lon_r"] = _round_coord(pd.to_numeric(work["lon"], errors="coerce"))
    work["_name_norm"] = work["plant_name"].astype(str).str.strip().str.lower()

    records = []
    for (country, _name_norm, _lat_r, _lon_r), group in work.groupby(
        ["country", "_name_norm", "_lat_r", "_lon_r"], dropna=False
    ):
        capacity = pd.to_numeric(group["capacity_mw"], errors="coerce")
        fuel_values = sorted(set(group["fuel_type"].dropna().astype(str).str.strip()))
        mixed = len(fuel_values) > 1
        years = pd.to_numeric(group["commissioning_year"], errors="coerce").dropna()

        records.append(
            {
                "country": country,
                "plant_name": group["plant_name"].iloc[0],
                "lat": pd.to_numeric(group["lat"], errors="coerce").mean(),
                "lon": pd.to_numeric(group["lon"], errors="coerce").mean(),
                "capacity_mw": capacity.sum(min_count=1),
                "fuel_type": None if mixed else (fuel_values[0] if fuel_values else None),
                "mixed_fuel_type": mixed,
                "fuel_types_found": ";".join(fuel_values),
                "commissioning_year": years.min() if len(years) else np.nan,
                "n_units": len(group),
            }
        )
    return pd.DataFrame.from_records(records)


def add_fuel_bucket(df: pd.DataFrame) -> pd.DataFrame:
    """Add ``fuel_type_bucket`` in {hydro, wind, solar, thermal} without
    touching ``fuel_type``/``fuel_types_found``. Raise on any unmapped fuel
    type or unlisted mixed plant."""
    df = df.copy()

    def _bucket(row) -> str:
        if row["mixed_fuel_type"]:
            override = MIXED_FUEL_BUCKET_OVERRIDES.get(row["plant_name"])
            if override is None:
                raise ValueError(
                    f"Mixed-fuel plant without an override: '{row['plant_name']}' "
                    f"(fuel_types_found={row['fuel_types_found']!r}). Add it to "
                    f"MIXED_FUEL_BUCKET_OVERRIDES."
                )
            return override
        bucket = FUEL_TYPE_TO_BUCKET.get(row["fuel_type"])
        if bucket is None:
            raise ValueError(
                f"fuel_type '{row['fuel_type']}' (plant '{row['plant_name']}') is not "
                f"in FUEL_TYPE_TO_BUCKET."
            )
        return bucket

    df["fuel_type_bucket"] = df.apply(_bucket, axis=1)
    return df


def fuel_bucket_distribution(countries: list[str] | None = None) -> pd.DataFrame:
    """Plant count and summed capacity per ``fuel_type_bucket`` per country,
    from the already-written ``gem_validated_plants_{country}.csv`` files."""
    countries = countries or COUNTRIES
    frames = []
    for country in countries:
        path = ASSETS_PROCESSED / f"gem_validated_plants_{country}.csv"
        if not path.exists():
            raise FileNotFoundError(f"{path} does not exist. Run validate() first.")
        df = pd.read_csv(path)
        grouped = (
            df.groupby("fuel_type_bucket")
            .agg(
                n_plants=("plant_name", "count"),
                capacity_mw_sum=("capacity_mw", lambda s: s.sum(min_count=1)),
            )
            .reset_index()
        )
        grouped.insert(0, "country", country)
        frames.append(grouped)
    result = pd.concat(frames, ignore_index=True)
    result["capacity_mw_sum"] = result["capacity_mw_sum"].round(1)
    return result.sort_values(["country", "capacity_mw_sum"], ascending=[True, False]).reset_index(drop=True)


def validate(path: Path) -> dict:
    """Full pipeline — see the module docstring for the exact order."""
    df = _apply_mapping(read_asset_table(path))
    operating, planned, status_summary = split_by_status(df)

    ASSETS_PROCESSED.mkdir(parents=True, exist_ok=True)

    planned_path = ASSETS_PROCESSED / "gem_planned_assets.csv"
    planned.to_csv(planned_path, index=False)

    operating_scope = operating[
        operating["country"].astype(str).str.strip().isin(COUNTRIES)
    ].copy()

    units_detail_path = ASSETS_PROCESSED / "gem_units_detail.csv"
    operating_scope.to_csv(units_detail_path, index=False)

    aggregated = add_fuel_bucket(aggregate_by_plant(operating_scope))

    report = {
        "source_file": str(path),
        "status_summary": status_summary,
        "planned_assets_path": str(planned_path),
        "units_detail_path": str(units_detail_path),
        "duplicate_coord_tolerance_deg": DUPLICATE_COORD_TOLERANCE_DEG,
        "fuel_buckets": sorted(set(FUEL_TYPE_TO_BUCKET.values())),
        "countries": {},
    }

    for country in COUNTRIES:
        sub_units = operating_scope[
            operating_scope["country"].astype(str).str.strip() == country
        ]
        sub = aggregated[aggregated["country"].astype(str).str.strip() == country].copy()

        if sub.empty:
            report["countries"][country] = {
                "n_units_operating_before_aggregation": len(sub_units),
                "n_plants_after_aggregation": 0,
                "note": "no operating plants found — check the 'Country/area' value "
                        "against config.COUNTRIES.",
            }
            continue

        n_excluded_mainland = 0
        excl_path = None
        try:
            xmin, ymin, xmax, ymax = get_country_bounds(country)
            lat = pd.to_numeric(sub["lat"], errors="coerce")
            lon = pd.to_numeric(sub["lon"], errors="coerce")
            in_bbox = (lat.between(ymin, ymax) & lon.between(xmin, xmax)).fillna(False)
            n_out_of_bbox = int((~in_bbox).sum())

            if country in MAINLAND_ONLY_COUNTRIES and n_out_of_bbox > 0:
                if country not in MAINLAND_EXCLUSION_FILENAMES:
                    raise NotImplementedError(
                        f"{country} is in MAINLAND_ONLY_COUNTRIES but has no entry in "
                        f"MAINLAND_EXCLUSION_FILENAMES."
                    )
                excluded = sub[~in_bbox].copy()
                excluded["exclusion_reason"] = (
                    f"outside the mainland bounding box of {country} "
                    f"(MAINLAND_ONLY_COUNTRIES) — likely Azores / Madeira."
                )
                excl_path = ASSETS_PROCESSED / MAINLAND_EXCLUSION_FILENAMES[country]
                excluded.to_csv(excl_path, index=False)
                n_excluded_mainland = len(excluded)
                sub = sub[in_bbox].copy()
                logger.info(
                    "%s: removed %d plant(s) outside the mainland bbox -> %s",
                    country, n_excluded_mainland, excl_path,
                )
        except FileNotFoundError as exc:
            n_out_of_bbox = None
            logger.warning("%s: GADM boundary absent, skipping bbox check (%s)", country, exc)

        capacity = pd.to_numeric(sub["capacity_mw"], errors="coerce")
        out_path = ASSETS_PROCESSED / f"gem_validated_plants_{country}.csv"
        sub.to_csv(out_path, index=False)

        report["countries"][country] = {
            "n_units_operating_before_aggregation": len(sub_units),
            "n_plants_after_aggregation": len(sub),
            "n_missing_or_zero_capacity": int((capacity.isna() | (capacity == 0)).sum()),
            "n_out_of_country_bbox": n_out_of_bbox,
            "n_excluded_mainland_only": n_excluded_mainland,
            "excluded_mainland_path": str(excl_path) if excl_path else None,
            "n_mixed_fuel_type": int(sub["mixed_fuel_type"].sum()),
            "bucket_counts": sub["fuel_type_bucket"].value_counts().to_dict(),
            "output_path": str(out_path),
        }

    return report


def print_report(report: dict) -> None:
    print(f"GEM validation — source: {report['source_file']}")
    ss = report["status_summary"]
    print(f"  operating={ss['n_operating']} planned={ss['n_planned']} other={ss['n_other_excluded_from_both']}")
    for country, stats in report["countries"].items():
        print(f"\n{country}:")
        if stats.get("n_plants_after_aggregation", 0) == 0 and "note" in stats:
            print(f"  {stats['note']}")
            continue
        print(f"  units (pre-aggregation): {stats['n_units_operating_before_aggregation']}")
        print(f"  plants (post-aggregation): {stats['n_plants_after_aggregation']}")
        print(f"  excluded (mainland-only): {stats['n_excluded_mainland_only']}")
        print(f"  out of country bbox: {stats['n_out_of_country_bbox']}")
        print(f"  mixed fuel type: {stats['n_mixed_fuel_type']}")
        print(f"  buckets: {stats['bucket_counts']}")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--discover", action="store_true", help="Report columns/types/sample.")
    mode.add_argument("--validate", action="store_true", help="Run the full pipeline.")
    mode.add_argument("--fuel-distribution", action="store_true", help="Bucket distribution from validated CSVs.")
    args = parser.parse_args()

    if args.fuel_distribution:
        try:
            dist = fuel_bucket_distribution()
        except FileNotFoundError as exc:
            logger.error(str(exc))
            sys.exit(1)
        OUTPUT_INSPECTION.mkdir(parents=True, exist_ok=True)
        dist.to_csv(OUTPUT_INSPECTION / "fuel_bucket_distribution.csv", index=False)
        print(dist.to_string(index=False))
        sys.exit(0)

    try:
        gem_file = find_gem_file()
    except FileNotFoundError as exc:
        logger.error(str(exc))
        sys.exit(1)

    if args.discover:
        discover_schema(gem_file)
        sys.exit(0)

    try:
        report = validate(gem_file)
    except (ValueError, NotImplementedError) as exc:
        logger.error(str(exc))
        sys.exit(1)

    print_report(report)
    OUTPUT_INSPECTION.mkdir(parents=True, exist_ok=True)
    report_path = OUTPUT_INSPECTION / "gem_validation_report.json"
    report_path.write_text(json.dumps(report, indent=2, default=str))
    logger.info("Report saved to %s", report_path)


if __name__ == "__main__":
    main()
